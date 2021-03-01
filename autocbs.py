#!/usr/bin/env python
#  Copyright (c) 2021 Freyr Yggdrasil 
#  https://github.com/FreyrYggdrasil/autocbs

#  Permission is hereby granted, free of charge, to any person
#  obtaining a copy of this software and associated documentation
#  files (the "Software"), to deal in the Software without
#  restriction, including without limitation the rights to use,
#  copy, modify, merge, publish, distribute, sublicense, and/or sell
#  copies of the Software, and to permit persons to whom the
#  Software is furnished to do so, subject to the following
#  conditions:

#  The above copyright notice and this permission notice shall be
#  included in all copies or substantial portions of the Software.

#  THE SOFTWARE IS PROVIDED "AS IS", WITHOUT WARRANTY OF ANY KIND,
#  EXPRESS OR IMPLIED, INCLUDING BUT NOT LIMITED TO THE WARRANTIES
#  OF MERCHANTABILITY, FITNESS FOR A PARTICULAR PURPOSE AND
#  NONINFRINGEMENT. IN NO EVENT SHALL THE AUTHORS OR COPYRIGHT
#  HOLDERS BE LIABLE FOR ANY CLAIM, DAMAGES OR OTHER LIABILITY,
#  WHETHER IN AN ACTION OF CONTRACT, TORT OR OTHERWISE, ARISING
#  FROM, OUT OF OR IN CONNECTION WITH THE SOFTWARE OR THE USE OR
#  OTHER DEALINGS IN THE SOFTWARE.

#  main import modules
import json
import cbsodata

#  **************************************************
#  from typing import List
import pickle
import itertools as it

#  format output file and file name and running time
import datetime 
import csv
import os
import glob
import textwrap

#  dataframes and manipulation
#  doc: https://pandas.pydata.org/pandas-docs/stable/reference/api/
#       pandas.DataFrame.to_excel.html
import pandas as pd 
import numpy as np
import functools
import operator
from pathlib import Path
import re

#  excel support
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Alignment
from openpyxl.utils.cell import coordinate_from_string
from openpyxl.utils.cell import column_index_from_string
from openpyxl.utils.cell import get_column_interval
from openpyxl.utils.cell import cols_from_range

#  command line arguments
import sys
import argparse

#**********************************
__version__ = '0.1.1'

 
def get_defaults():
    """
    Helper method for getting the default settings.

    Returns
    -------
    default_settings : dict
        A dictionary of the default settings.
    """

    return {
        "help_switch": False,
        "outputfolder": "",
        "data_refresh": False,
        "table_meta": False,
        "download_data": False,
        "no_master": False,
        "masterConvertToHtml":False,
        "download_excel": False,
        "search_arg": [],
        "table_prop": False,
        "download_csv": False,
        "download_json": False,
        "force_download": False,
        "modified_within": "",
        "table_ids": [],
        "search_regex": [],
        "get_tables": 0,
        "loglevel": "info",
        "start_record": 0,
    }

# **************************************************
# print string to screen for user feedback
def p(plevel:int, text, *args) -> bool:

    global settings
    global loglevels

    level = settings["loglevel"]
    print_line = False
    no_linefeed = False

    try:
        if loglevels.index(settings["loglevel"]) >= plevel: 
            print_line = True
    except Exception as e:
        pass

    if print_line:
        if not text: 
            text = ''
        elif type(text) == type(list()):
            # no lists
            text = ''
        else:
            text = re.sub(' +', ' ', str(text))

        try:
            if args:
                for i in args:
                    if not i == 'end=""':
                        text = text + ' ' + str(i)
                    else:
                        no_linefeed = True
        except Exception as f:
            pass

        if no_linefeed:
            print(text, end="")
        else:
            print(text)

    return True

# **************************
# excel helpers
def transpose(ws, min_row: int, max_row: int, min_col: int, 
    max_col: int):
    
    for row in range(min_row, max_row+1):
        for col in range(min_col, max_col+1):
            ws.cell(row=col,column=row).value = \
            ws.cell(row=row,column=col).value

def transpose_row_to_col(ws, 
        min_row: int, max_row: int, 
        min_col: int, max_col: int, 
        target_cell_address=(1,1), delete_source=False):
        
    cell_values = []
    for row in ws.iter_rows(min_row=min_row, max_row=max_row, 
                            min_col=min_col, max_col=max_col):
        for cell in row:
            cell_values.append(cell.value)
            if delete_source:
                cell.value = ""
    fill_cells(ws, target_cell_address[0], target_cell_address[1], 
                   cell_values)

def transpose_col_to_row(ws, min_row, max_row, min_col, max_col, 
                         target_cell_address=(1,1), delete_source=False):
    cell_values = []
    for col in ws.iter_cols(min_row=min_row, max_row=max_row, 
                            min_col=min_col, max_col=max_col):
        for cell in col:
            cell_values.append(cell.value)
            if delete_source:
                cell.value = ""
    fill_cells(ws, target_cell_address[0], target_cell_address[1], 
                   cell_values)

def fill_cells(ws, start_row, start_column, cell_values):
    row = start_row
    column = start_column
    for value in cell_values:
        ws.cell(row=row,column=column).value = value
        row += 1

def convertTuple(tup): 
    str = functools.reduce(operator.add, (tup)) 
    return str

# end excel helpers      
     
# *********************************
# save objects
def save_data(data, dir, p_identifier, metadata_name, argument):

    if type(argument) == type(str()):
        output_file = os.path.join(dir, p_identifier+'-'+metadata_name 
                                        + '.' + argument)
        
    elif type(argument) == type(None):
        p(critical,'Argument for function is empty', type(argument))
        raise SystemExit(1)
    else:
        # getting data for excel
        output_file = os.path.join(dir, p_identifier+'-objects.xlsx')
        workbook = argument

    if argument == 'json':
        my_data = json.loads(str(data))
        with open(output_file, 'w') as output_file:
            json.dump(my_data, output_file, indent=4)
        output_file.close()
        
        # update date mc
        if not settings["no_master"]:
            controlInformationTable['lastRefreshDateJson'] = \
            datetime.datetime.now().strftime("%Y-%m-%dT%H:%M:%S")
        
        return my_data
        
    elif argument == 'csv':
        data_csv = data.to_csv(output_file, sep=";", na_rep="",
                               quoting=csv.QUOTE_ALL, quotechar='"',
                               doublequote=True, escapechar="\\", 
                               index = False)
        
        # update date mc
        if not settings["no_master"]:
            controlInformationTable['lastRefreshDateCsv'] = 
            datetime.datetime.now().strftime("%Y-%m-%dT%H:%M:%S")

        return data_csv
    
    else:
        # excel sheet data
        if str(type(workbook)) != 
                "<class 'openpyxl.workbook.workbook.Workbook'>" and 
                type(workbook) != None:
            workbook = convertTuple(workbook)
        
        try:
            sheet = workbook[metadata_name[0:30]]
        except:
            p(info,'\t\t\t\t... creating', metadata_name[0:30], 
                   'in workbook.')
            sheet = workbook.create_sheet(metadata_name[0:30])
        
        for row in dataframe_to_rows(data, index=False, header=True):
            sheet.append(row)

        if metadata_name == 'TableInfos' or metadata_name == 'TableListInfo':
            # transpose TableInfos for easier reading
            start, stop = 1, sheet.max_column
            transpose(sheet, min_row=1, max_row=1, min_col=1, 
                             max_col=sheet.max_column)
            transpose_row_to_col(sheet, min_row=1, max_row=1, 
                                        min_col=1, max_col=sheet.max_column,
                                        target_cell_address=(3,1))
            transpose(sheet, min_row=2, max_row=2, min_col=1, 
                             max_col=sheet.max_column)
            transpose_row_to_col(sheet, min_row=2, max_row=2, 
                                        min_col=1, max_col=sheet.max_column,
                                        target_cell_address=(3,2))
            sheet.delete_rows(1,2)
            sheet.column_dimensions['A'].width = 16
            sheet.column_dimensions['B'].width = 100            
            for index, row in enumerate(sheet.iter_rows()):
                if start < index < stop:
                    for cell in row:
                        cell.alignment = Alignment(horizontal='left', 
                                                   vertical='center', 
                                                   wrap_text=True)

        return workbook 
        
# **************************
# master control data
def masterControlData(data):
    # data object is TableInfos
    global controlInformationTables # all Tables evaluated
    global controlInformationTable  # current Table evaluated

    if not settings["no_master"]:
    
        for ci in controlInformationTable:
            try:
                controlInformationTable[ci] = data[ci]
            except KeyError:    # first key for extra autocbs values
                controlInformationTable['lastRefreshDate'] = 
                    datetime.datetime.now().strftime("%Y-%m-%dT%H:%M:%S")
                break
                
    return
    
# ********************************************
# save master control file xlsx
# uses global var 
# controlInformationTable
# settings["outputfolder"]
# control_file
def masterControlFile(*fromProp):
    global controlInformationTable
    global controlInformationTables
    
    output_file = settings["outputfolder"] + control_file
    
    if not fromProp:
            
        if Path(output_file).is_file():            
            try:
                controlbook = load_workbook(output_file)
                controlbook.active = controlbook['controlbook CBS']
                sheet = controlbook.active
                addHeader = False
                addValues = True

            except:
                controlbook = Workbook()    
                sheet = controlbook.create_sheet('controlbook CBS')
                addHeader = True
                addValues = True
                
        else:
            controlbook = Workbook()    
            sheet = controlbook.create_sheet('controlbook CBS')
            addHeader = True
            addValues = True
            
        lastrow = sheet.max_row+1
        
        if addHeader:

            for col, val in enumerate(controlInformationTables.keys(), 
                            start=1):
                sheet.cell(row=1, column=col).value = val.encode("utf8")
                            
        if addValues:
            for col, val in enumerate(controlInformationTables.values(), 
                            start=1):
                if type(val) != type(int):
                    sheet.cell(lastrow, column=col).value = val
                else:
                    sheet.cell(lastrow, column=col).value = 
                                                    str(val).encode("utf8")
                    
        try:
            controlbook.save(output_file) 
            
        except Exception as e:
            p(warning,'\nUnable to save master control workbook', 
                       output_file, "Do you have it open in Excel? \
                       The error message is", e)
            pass    

        controlInformationTable = {"Title":"","Updated":"",
                                   "ShortTitle":"","Identifier":"",
                                   "Summary":"","Modified":"",
                                   "ReasonDelivery":"","Frequency":"",
                                   "Period":"","RecordCount":"",
                                   "lastRefreshDate":"",
                                   "lastRefreshDateJson":"",
                                   "lastRefreshDateCsv":"",
                                   "lastRefreshDateExcel":"",
                                   "statLineURL":""}
        controlInformationTables = {}

    elif convertTuple(fromProp) == 'html' or 
         convertTuple(fromProp) == 'remdup':
        
        if convertTuple(fromProp) == 'html': 
            p(info, '\nConverting master controlbook '+output_file+
                    ' to html file '+output_file+'.html')
        
        if Path(output_file).is_file():
            try:
                controlbook = load_workbook(output_file, data_only = True)
                sheet_names = controlbook.sheetnames
            except:
                p(info,'Unable to load controlbook', output_file, 
                       'the error was\n', e)
                sheet_names = ['controlbook CBS']
                pass
                
            try:
                data = {}
                with pd.ExcelFile(output_file) as xls:
                    for sh_name in sheet_names:
                        data[sh_name] = pd.read_excel(xls, sh_name, 
                                                      index_col=None)
                
                if convertTuple(fromProp) == 'remdup':
                    try:
                        data[sh_name] = data[sh_name].
                                        drop_duplicates(subset=['Identifier']
                                        , keep='last')
                        with pd.ExcelWriter(output_file) as writer:
                            for sh_name in sheet_names:
                                try:
                                    data[sh_name].to_excel(writer, 
                                            sheet_name=sh_name, index=False)
                                except Exception as e:
                                    p(error, 'An error occured wil writing \
                                            the deduplicated file:', e)
                    
                    except Exception as e:
                        p(info,'Unable to remove duplicates from controlbook'
                                , output_file, '\nThe error was: ', e)
                        return
                
            except Exception as e:
                p(info,'Unable to load panda controlbook', output_file, 
                        'the error was\n', e)
                return
            
            finally:
                
                try:
                    controlbook = load_workbook(output_file, 
                                                data_only = True)    
                except Exception as e:
                    p(info,'Unable to load deduplicated controlbook', 
                            output_file, 'the error was\n', e)
                    return

        html_data = """<!DOCTYPE html>
<html lang="nl">
<head><title>controlbook CBS data</title>
<style>
body {
  background-color: white;
}
h1 {
  color: blue;
  font-family: verdana;
  font-size: 300%;
}
h2 {
  font-family: verdana;
  font-size: 200%;
}
h3 {
  color: blue;
  font-family: verdana;
  font-size: 150%;    
}
body {
  color: black;
  font-family: verdana;
  font-size: 100%;    
}
table {
  border: 1px solid black;
  border-collapse: collapse;
  width: 100%;
  margin-left: 10px;
  margin-right: 10px;
  color: black;
  font-family: verdana;
  font-size: 100%;   
}
th, td.header {
  border: 1px solid black;
  border-collapse: collapse;
  height: 20px;
  text-align: left;
  background-color: #4CAF50;
  color: white;  
}
td, cellhref, cellempty, cellvalue {
  border: 1px solid black;
  border-collapse: collapse;
  text-align: left;
  vertical-align: top;
}
th, td {
  padding: 8px;
}
tr {
  border: 1px solid black;
  border-collapse: collapse;
}
tr:hover {
  background-color: #f5f5f5;
}
div {
  height:60vh;
  overflow-y:scroll;
  overflow-x:scroll;
}
</style>
</head>
<body>
<h3>controlbook CBS data</h3>
<div>
<table>"""

        ws_sheet = controlbook['controlbook CBS']
        a = 0
        statLineURLs = []

        for cell in ws_sheet['O']:
            statLineURLs.append(cell.value)

        end_column = ws_sheet.max_column
        max_rows = ws_sheet.max_row
        start_column = 1

        for row in ws_sheet.iter_rows(
                            min_row=1, 
                            max_col=end_column, 
                            max_row=max_rows):
            if a == 0: # first row
                html_data += '\n<tr class="header">' 
            else: # next rows
                html_data += '\n<tr>'

            column_index = start_column        

            for cell in row:

                if a == 0:
                    html_data += '\n\t<th class="header">' \
                                 + str(cell.value) + '</th>'

                if a > 0:  # next rows
                    if cell.value is None:
                        html_data += '\n\t<td class="cellempty"> </td>'

                    else:
                        if column_index == 1:   # first column
                            html_data += '\n\t<td class="celhref"> \
                                <a href="' + statLineURLs[a] \
                                + '" title="'  + 'Naar StatLine' \
                                + '">'+ str(cell.value) + '</a></td>' 
                        else:
                            html_data += '\n\t<td class="cellvalue">' \
                                         + str(cell.value) + '</td>'
                
                column_index+=1 # next column                
                
                if column_index == end_column:
                    html_data += "</tr>"
                    a+=1
                    break

 
        html_data += "</table></div></body></html>"

        with open(output_file+'.html', "w") as html_fil:
            html_fil.write(html_data)

        html_fil.close()

    return

# ****************************
# get endpoint using cbsodata
def get_table_endpoint(p_identifier, endpoint, file_path, workbook):

    if endpoint == 'TableListInfo':
        data = p_identifier
        p_identifier = data['Identifier']
    else:
        p_identifier = p_identifier

    file_exists = False

    if str(type(workbook)) == \
            "<class 'openpyxl.workbook.workbook.Workbook'>" and \
            settings["download_excel"]:
        for sheet_title in workbook.sheetnames:
            if sheet_title == endpoint[0:30]:
                sheet = workbook[sheet_title]
                break
            else:
                sheet = workbook[sheet_title]

    # does file exist and do we update and/or force?
    # otherwise skip
    if not settings["force_download"] and \
        not settings["data_refresh"]:
        if settings["download_csv"]:
            output_file = os.path.join(
                                file_path, 
                                p_identifier+'-'+endpoint+ '.csv'
                                )
            if Path(output_file).is_file():
                file_exists = True

        if settings["download_json"]:
           output_file = os.path.join(
                                file_path, 
                                p_identifier+'-'+endpoint+ '.json'
                                )
           if Path(output_file).is_file():
                file_exists = True

        if settings["download_excel"]:
            output_file = os.path.join(
                                file_path, 
                                p_identifier+'-objects.xlsx'
                                )
            if Path(output_file).is_file():
                file_exists = True

        if file_exists and not settings["download_excel"]: 
            p(info, '\t\t\t\t...\tfile exists, not updating')
            
    elif settings["force_download"] and settings["download_excel"]:
        settings["data_refresh"] = True
        output_file = os.path.join(file_path, p_identifier+'-objects.xlsx')
        if Path(output_file).is_file():
            file_exists = True       

    else:
        file_exists = False

    # do csv & json & excel
    if (not file_exists and settings["download_data"]) and (settings["download_csv"] or settings["download_json"] or settings["download_excel"]):

        try:
            if endpoint == 'TableListInfo':
                # put tablelistinfo in dataframe
                data = pd.DataFrame(data, index=[0])
            else:
                # put cbsodata respons in dataframe
                data = pd.DataFrame(
                        cbsodata.get_meta(p_identifier, endpoint))
            
        except Exception as e:
            p(warning, '\t\t\t\tUnable to retrieve object', 
                        endpoint, 
                        'for table', 
                        p_identifier, 
                        '. The error message was\t\t\t\t', 
                        e)
            return

        if settings["download_excel"] and workbook:
            if type(workbook) != None: 
                if str(type(workbook)) != \
                    "<class 'openpyxl.workbook.workbook.Workbook'>":
                    workbook = convertTuple(workbook)            

            try:
                # just remove this sheet, leave the rest alone
                workbook.remove(workbook[endpoint[0:30]])
                
            except Exception as e:
                # the sheet for this data was not found
                pass
                
            finally:
                # create sheet for data
                for sheet_title in workbook.sheetnames:
                    sheet = workbook[sheet_title]
                else:
                    sheet = workbook.create_sheet(endpoint[0:30])
                workbook.active = workbook[endpoint[0:30]]
                sheet = workbook.active 

        if endpoint == 'DataProperties':
            # get extra endpoints
            # will be added to excel file as sheets
            # and .csv/.json files on disk
            data_np = data[['odata.type', 'Key']].to_numpy()
            for dimension in data_np:
                if dimension[0] == 'Cbs.OData.Dimension':
                    p(info, '\t\t\t\t\t... extra Dimension', dimension[1])
                    # do it again
                    workbook = get_table_endpoint(p_identifier, 
                                        dimension[1], file_path, workbook) 
            
            for period in data_np:
                if period[0] == 'Cbs.OData.TimeDimension':
                    p(info, '\t\t\t\t\t... extra TimeDimension', period[1])
                    # do it again
                    workbook = get_table_endpoint(p_identifier, 
                                            period[1], file_path, workbook) 
            
        if settings["download_csv"]:
            csv_data = save_data(data, file_path, 
                                p_identifier, endpoint,'csv')
            return csv_data
            
        if settings["download_json"]:
            json_data = save_data(data.to_json(), file_path, 
                                        p_identifier, endpoint, 'json')
            return json_data
            
        if settings["download_excel"]:
            workbook = save_data(data, file_path, 
                            p_identifier, endpoint, workbook)
            return workbook
            
    elif file_exists and (settings["download_data"] or 
                          settings["data_refresh"]) and 
                          settings["download_excel"]:

        try:                
            if endpoint == 'TableListInfo':
                # put tablelistinfo in dataframe
                data = pd.DataFrame(data, index=[0])
            else:
                # put cbsodata respons in dataframe
                data = pd.DataFrame(cbsodata.get_meta(p_identifier,
                                    endpoint))
            
        except Exception as e:
            p(warning, '\t\t\t\tUnable to retrieve object', 
                        endpoint, 'for table', p_identifier, 
                        '. The error message was\t\t\t\t', e)
            return
                    
        workbook = save_data(data, file_path, 
                             p_identifier, endpoint, workbook)
        
        return workbook
    
    return

#  ********
#  check if dir exists
#  if not create it
#  > returns True|False
def checkDirCreate(folderpath: str) -> bool:
    if not os.path.isdir(folderpath):
        try:
            os.mkdir(folderpath)
            return True
        except Exception as e:
            p(error, 'Creating folder', folderpath
                     , 'failed with error', e
                     , 'Do you have sufficient rights?'
             )
            return False
    else: return True
    
#  ********
#  download data from table
#  > return workbook
def get_table_meta(data, endpoint, workbook):
    
    p_identifier = data['Identifier']
    
    file_path = settings["outputfolder"]+p_identifier+'/'

    settings["download_data"] = checkDirCreate(file_path)

    if settings["download_excel"]:
        output_file = file_path+p_identifier+"-objects.xlsx"

        if Path(output_file).is_file():
            # file exists, read it
            if settings["data_refresh"]:
                try:
                    # does the workbook exist? if so update sheets.
                    # don't mess with other sheets
                    workbook = load_workbook(
                                file_path+p_identifier+"-objects.xlsx")
                except:
                    # create new 
                    workbook = Workbook()
                    
            elif settings["force_download"]: 
                # create new by -force
                workbook = Workbook()  
                
        else:
            workbook = Workbook()    
    else:
        workbook = workbook
        
    # get the data endpoints from list
    p(info, '\t\t\t\t... performing update for', endpoint)
    if endpoint == 'TableListInfo':
        workbook = get_table_endpoint(data, endpoint, file_path, workbook)
    else:
        if settings["download_excel"] or settings["download_csv"] 
                or settings["download_json"]:
            workbook = get_table_endpoint(p_identifier, 
                                    endpoint, file_path, workbook)

    # saving excel
    if settings["download_excel"] and str(type(workbook)) == 
                        "<class 'openpyxl.workbook.workbook.Workbook'>":
        if not settings["no_master"]:
            controlInformationTable['lastRefreshDateExcel'] = 
                        datetime.datetime.now().strftime("%Y-%m-%dT%H:%M:%S")

        try:
            workbook.save(file_path+p_identifier+"-objects.xlsx")

        except AttributeError as e:
            p(error, 'an AttributeError appeared while trying to \
                      save the excel file\n', e)
            if not settings["data_refresh"] and not 
                   settings["force_download"]:
                pass
                
        except Exception as e:
            p(warning,'\nUnable to save workbook', 
                         file_path+p_identifier+"-objects.xlsx.", 
                         "Do you have it open in Excel?\n \
                         The error message is \n", e)
            pass

                    
    return workbook

#  ********
#  Convert datatype to list 
#  > return listObject
def convertToList(data, datatype):
    
    try:
        if str(datatype) == 'CBSODATA':
            #  create local copy of table list
            listObject = []
            try:
                with open(data, "rb") as f:
                    listObject = pickle.load(f) 
                f.close()                
            except Exception as e:
                listObject = []
                print(e)
            return listObject
            
        elif str(datatype) == 'CONTROL':
            #  split for excel
            listObject = []
            try:
                listObject = data.split(',')                
            except Exception as e:
                listObject = []
                print(e)
            return listObject

        else:
            #  others might follow
            p(warning, 'Converting the file', file, 'of type', 
                        datatype, 'did not succeed. It seems like', 
                        datatype, 'is not implemented (yet).')
            listObject = []
            return listObject
            
    except Exception as e:
        p(error, 'While converting the file', file, 'to a list the \
                  following error occured:\n', e)
        listObject = []
        return listObject

#  ********
#  evaluate
def startEvaluating():
    
    # some vars for loop
    all_tables = len(tables)    
    end_record = all_tables
    itable = 0   # count tables processed
    itable_records = 0  # count nr of records in hits
    number_of_hits = 0  # count search hits
    result_list = []    # save table identifers

    if not settings["start_record"]: 
        settings["start_record"] = 0
    else:
        p(verbose, 'Starting at record ', settings["start_record"])

    p(verbose, 'Table list contains ' + str(all_tables) + ' tables ', 
               'starting at record '+ str(settings["start_record"]) 
               if settings["start_record"] > 0 else '' )

    # how much?
    end_record = all_tables - settings["start_record"]
        
    if settings["get_tables"] > 0: 
        p(info, 'Getting maximum of ' + str(settings["get_tables"]) + 
                ' tables due to argument -n or -i.')
        end_record = settings["start_record"] + settings["get_tables"]
        
    if end_record > all_tables:
        end_record = all_tables
        p(verbose, 'End record is ' + str(end_record))
        
    # Console messages 
    p(info, 'Searching in ShortDescription for keyword(s) ' + 
             str(settings["search_arg"]) if settings["search_arg"] 
             else 'No search keywords (-s) given.')
    p(info, 'Downloading data into folder ', settings["outputfolder"])

    if settings["table_meta"]: p(verbose, 'Meta data will be downloaded...')

    # modified date
    if settings["modified_within"]:
        if settings["modified_within"] == 'lastday':
            minmoddate = datetime.date.today() - datetime.timedelta(days=1)
        elif settings["modified_within"] == 'lastweek':
            minmoddate = datetime.date.today() - datetime.timedelta(days=7)
        elif settings["modified_within"] == 'lastmonth':
            minmoddate = datetime.date.today() - datetime.timedelta(days=30)
        elif settings["modified_within"] == 'lastyear':
            minmoddate = datetime.date.today() - datetime.timedelta(days=365)
        else:
            try:
                minmoddate = datetime.datetime.strptime(
                             settings["modified_within"], '%Y-%m-%d')
                minmoddate = datetime.datetime.date(minmoddate)
            except ValueError:
                try:
                    minmoddate = datetime.datetime.strptime(
                                 settings["modified_within"], '%Y%m%d')
                    minmoddate = datetime.datetime.date(minmoddate)
                except ValueError:
                    p(error,'Can\'t understand the modified date? \
                             It should be in the form 20210101?', 
                             settings["modified_within"])
                    settings["modified_within"] = False
                    raise SystemExit(16)
                    
        maxmoddate = datetime.date.today()
        p(info,'Looking for modifications between', 
                minmoddate, 'and', maxmoddate, '\n')

    # start loop for all tables
    p(verbose, '\n--------------------')

    for table in tables:
        itable+=1
        if settings["table_ids"]:
            table=table[0]  
        # loop until we get at the start record
            
        if itable >= settings["start_record"] and itable <= end_record:  
            
            if settings["modified_within"]:
                datemodified = datetime.datetime.strptime(
                               table['Modified'][0:10], '%Y-%m-%d')
                datemodified = datetime.datetime.date(datemodified)
                if datemodified >= minmoddate and datemodified <= maxmoddate:
                    modified_within_selection = True
                    p(verbose, 'Table', table['Identifier'], 
                               'modified on', datemodified, 
                               'which is valid for the selection.')
                elif datemodified > maxmoddate:
                    modified_within_selection = False
                    p(verbose, 'Table has a modified date of', 
                                datemodified, 'which is in the future. Wauw.')
                elif datemodified < minmoddate: 
                    modified_within_selection = False
                    p(verbose, 'Table is last modified on', 
                                datemodified, 'which is too far in the past.')
            else:
                modified_within_selection = True
            
            if modified_within_selection:
                p(allmsg, '\nUsing meta data', table)
                p(allmsg, '\nIdentifier table ' + 
                            table['Identifier'] + '\nShortTitle table' + 
                            table['ShortTitle'])
                p(allmsg, '\nShortDescription table' + 
                             table['ShortDescription'])            

                # search properties
                if len(settings["search_keywords"]) > 0:
                    for keyword in settings["search_keywords"]:
                        isHit = False   # only one hit is needed
                        if not isHit:
                            if str(
                                table['ShortDescription']).find(keyword) > 0:
                                number_of_hits += 1
                                isHit = True
                                p(info if not 
                                    settings["download_data"] else 
                                    verbose, 
                                    table['Identifier'], 
                                    'has in ShortDescription search item', 
                                    keyword,'and is \nupdated on ' + 
                                    str(datemodified) + 
                                    ' which is in the modified date period.' 
                                    if modified_within_selection 
                                    else keyword)
                                result_list.append(table)
                            else:
                                p(verbose, 'Search keyword', 
                                    keyword, 'not found in table', 
                                    table['Identifier'])
                    
                else: # no search parameters given
                    p(verbose, 'Table', table['Identifier'], 
                               'selected and added to the result list.')
                    result_list.append(table)

        # stop searching
        if itable >= end_record:
            p(verbose, '--------------------')
            break

    # any results? or just this one table
    if len(result_list)>0:
        p(info, '\nNumber of tables to retrieve:', len(result_list))
        
        if len(result_list) > 60 and (not settings["data_refresh"] 
                                 or not settings["force_download"]) 
                                 and settings["download_data"] 
                                 and not settings["table_prop"]:
            p(warning, "\nThis is (probably) a lot of data, please \
                          use -force -update to download. \nOr use \
                          parameter -m to download just the table information.")
            
        else:
            for result in result_list:
                itable_records = itable_records + int(result['RecordCount'])
            
                if settings["download_data"]:

                    p(info, '\n\tCommencing retrievel of', 
                             result['RecordCount'],'records for table', 
                             result['Identifier'], 
                             result_list.index(result)+1,
                             '/',len(result_list))
                    
                    # initialize excel
                    workbook = Workbook()                
                    
                    if settings["table_prop"]:
                        get_table_meta(result, 'DataProperties', workbook)
                        # get info from tables
               
                    if settings["table_meta"]:
                        get_table_meta(result, 'TableInfos', workbook)
                        get_table_meta(result, 'TableListInfo', workbook)
                            
                    elif not settings["table_meta"] and 
                        not settings["table_prop"]:
                        # get all data and properties of table
                        objects_lst = ['DataProperties','TableInfos',
                                        'CategoryGroups','TypedDataSet',
                                        'TableListInfo']
                        
                        if settings["download_excel"]:
                            p(info,'\t\texcel file name '+
                                    result['Identifier']+'/'+
                                    result['Identifier']+
                                    "-objects.xlsx")
                        if settings["download_csv"]:
                            p(info,'\t\tcsv file name(s) ' +
                                    result['Identifier']+'/'+
                                    result['Identifier']+'-<object>.csv')
                        if settings["download_json"]:
                            p(info,'\t\tjson file name(s) ' +
                            result['Identifier']+'/'+
                            result['Identifier']+
                            '-<object>.json')

                        for object in objects_lst:
                            download_excel_old = settings["download_excel"]
                            if not int(result['RecordCount']) > 1000000 and 
                                not object == 'TypedDataSet':
                                get_table_meta(result, object, workbook)
                            elif int(result['RecordCount']) > 1000000 and 
                                object == 'TypedDataSet':
                                p(warning,'\t\t\t\t... data TypedDataset \
                                not converted to excel because\n\t\t\t\t... \
                                \tit has too many records ('+
                                str(result['RecordCount'])+
                                ').', 'Use -csv\n\t\t\t\t... \tto download \
                                csv file.' if not settings["download_csv"] 
                                else 'Next -csv\n\t\t\t\t... \tdownload \
                                    will occur.')
                                settings["download_excel"] = False
                                if settings["download_csv"] or 
                                    settings["download_json"]: 
                                    get_table_meta(result, object, workbook)
                                settings["download_excel"] = 
                                          download_excel_old
                            else:
                                get_table_meta(result, object, workbook)


                    if not settings["no_master"]:
                        masterControlData(result)
                        controlInformationTable['lastRefreshDate'] = 
                            datetime.datetime.now().strftime("%Y-%m-%dT%H:%M:%S")
                        controlInformationTable['statLineURL'] = 
                            'https://opendata.cbs.nl/#/CBS/nl/navigatieScherm/zoeken?searchKeywords=' + 
                            result['Identifier']
                        controlInformationTables.update(controlInformationTable)    
                        masterControlFile()        


            if not settings["download_data"]:
                p(warning, "\nResults not downloaded. Use argument '-d' \
                             for download or one\nof the file arguments \
                             (-c, -x, -j). When there are more\nthen 60 \
                             tables also use -force -update.")
                
            if settings["download_data"]:
                masterControlFile()
                if settings["masterConvertToHtml"]:
                    masterControlFile('remdup') # remove duplicates
                    masterControlFile('html') # create html output
        
        p(verbose, 'Finished retrieving results.')
        
    else:
        p(warning,'No search strings and/or results found. Maybe something \
            wrong with the filter?')

    # stats for geeks
    end = datetime.datetime.now() 
    elapsed_time = (end - begin)    
    p(silent, '\nTotal time passed ' + str(elapsed_time), 
              ' seconds, which is ', elapsed_time/len(result_list) 
              if len(result_list)>0 else 0, ' per\ntable. In total \
              there are', itable_records, '(unique) records in \
              the selection.' )

#  ********
#  Save CBS odata file as local copy
def pickleDumpTables(tables, output_file: str):
    try:
        pickle.dump(tables, output_file)
        return True
    except Exception as e:
        p(error, "The table list could not be saved to",
                    output_file, 
                    "The error messages was\n",
                    e)
        return False

#  ********
#  request list from odata endpoint
#  exits program when fails
def getTableListCBS(tables: list, *filter) -> list:
    try:
        if not filter:
            tables = cbsodata.get_table_list()
        else:
            tables = cbsodata.get_table_list(None, filter)
        return tables
        
    except Exception as e:
        if not filter:
            p(critical, "The table list could not be retrieved from \
                        the cbs odata endpoint. \
                        The error messages was\n",
                        e)
            raise SystemExit(1)
        else:
            p(warning, "The table [" + str(i) + "] could not be reached.",
                        "Are you using the correct Table Identifier?",
                        "The error message was\n", 
                        e)
            return []

#  ********
#  evaluate tables 
def getTableList():
    
    global all_tables
    global tables
    
    if not settings["table_ids"]:
        if len(glob.glob(
                settings["outputfolder"] 
                + "cbs_all_tables.list")) > 0 and \
                (settings["data_refresh"] or \
                settings["force_download"]):
            p(info, "A local copy of 'cbs_all_tables.json' has been found. \
                    Update is used, retrieving data from CBS odata endpoint \
                    and saving it as new local copy.")

            tables = getTableListCBS('all')
            # save as local copy
            if checkDirCreate(settings["outputfolder"]):
                success = pickleDumpTables(tables, 
                                settings["outputfolder"] 
                                + "cbs_all_tables.list")

        elif len(glob.glob(
                    settings["outputfolder"] 
                    + "cbs_all_tables.list")) > 0:
            # we have a cbs_all_tables local copy
            p(info, "Using local copy of CBS table information in \
                'cbs_all_tables.json'.")
            tables = convertToList(
                        settings["outputfolder"] 
                        + "cbs_all_tables.list", "CBSODATA")

        else:
            p(info, "No local copy of 'cbs_all_tables.list' was found. \
                    Retrieving data from CBS odata endpoint and saving \
                    it as new local copy.")

            tables = getTableListCBS()
            # save as local copy
            if checkDirCreate(settings["outputfolder"]):
                success = pickleDumpTables(tables, 
                                settings["outputfolder"] 
                                + "cbs_all_tables.list")

    else:
        # just this one table(s)
        tables = []
        p(info, "Retrieving data from CBS odata endpoint for \
                every table ID given.")

        for i in settings["table_ids"]:
            tables.append(getTableListCBS("Identifier eq '"+str(i)+"'"))
            p(info, "Table", 
                    i, 
                    "added to the selection list. Total", 
                    len(tables))

        if len(tables) == 0:
            p(critical,"A table identifier(s) was given, but none could \
                        be reached at the proper odata endpoint.")
            raise SystemExit(1)
        else:
            p(verbose, 'TableInfo downloaded from CBS for', 
                        len(tables), 
                        'due to argument -i.')
            settings["get_tables"] = len(tables)

#  ********
#  main function
def main():
 
    # doc: https://docs.python.org/3/library/argparse.html
    parser = argparse.ArgumentParser(
        prog="autocbs.py",
        usage=argparse.SUPPRESS,
        description="Automate CBS OData downloads "
                    "for ongoing projects. ",
        prefix_chars="-",
        fromfile_prefix_chars=None,
        add_help=True,
        exit_on_error=True,
        parents=[],
        allow_abbrev=False,
        argument_default=None,
        epilog="""
  following arguments use one or more parameter(s)
  --output, -o <path>         Give a path as the destination folder
  --search, -s <words>        Keywords to search for, seperated with space
  --identifiers, -i <id's>    Table identifier(s) to download. Multiple
                              identifers seperated with space
  --loglevel, -l <level>      Loglevel to use, one of 'silent', 'critical',
                              'error', 'warning', 'info', 'verbose' or 'allmsg'
  --modified, -w <key/date>   If used will check the modified date of the table
                              and check if it is valid for the period. The date
                              modified is between today and 'lastday', 'lastweek',
                              'lastmonth' or 'lastyear'. 
                              Alternatively a specific date can be given in the
                              format YYYY-MM-DD or YYYYMMDD
  
  When used without arguments the default is to update existing tables in the
  <current directory>+'/data' with loglevel 'critical'.""",
        formatter_class=argparse.RawDescriptionHelpFormatter,)
    parser.add_argument(
        "--download",
        "-d",
        dest="download_data",
        default=False,
        action="store_true",
        help="Download and save gathered (meta) data",
    )
    parser.add_argument(
        "--output",
        "-o",
        metavar='',
        dest="outputfolder",
        default="./data/",
        nargs="?",
        help=argparse.SUPPRESS,
    )
    parser.add_argument(
        "--search",
        "-s",
        type=str,
        metavar='',
        dest="search_keywords",     
        default="",
        nargs="+",
        help=argparse.SUPPRESS,
    )
    parser.add_argument(
        "--identifiers", 
        "-i",
        metavar='',
        type=str,
        dest="table_ids",
        default="",
        nargs="+",
        help=argparse.SUPPRESS,
    )
    parser.add_argument(
        "--returnmax",
        "-r",
        metavar='',
        type=int,
        dest="get_tables",
        default=0,
        nargs=1,
        help="Maximum numbers of tables to retrieve",
    )
    parser.add_argument(
        "--begin",
        "-b",
        metavar='',
        dest="start_record",
        default=0,
        nargs=1,
        help="Starts evaluating up from given record number",
    )
    parser.add_argument(
        "--loglevel",
        "-l",
        metavar='',
        dest="loglevel",
        default="verbose",
        choices=loglevels,
        nargs="?",
        help=argparse.SUPPRESS,
    )
    parser.add_argument(
        "--meta",
        "-m",
        dest="table_meta",
        default=False,
        action="store_true",
        help="Retrieve (only) table meta information",
    )
    parser.add_argument(
        "--nomaster",
        "-n",
        dest="no_master",
        default=False,
        action="store_true",        
        help="When used doesn\'t update the master excel file.",
    )
    parser.add_argument(
        "--convertHtml",
        dest="masterConvertToHtml",
        default=False,
        action="store_true",
        help="Remove duplicates from master file and export to html",
    )    
    parser.add_argument(
        "--properties",
        "-p",
        dest="table_prop",
        default=False,
        action="store_true",
        help="Retrieve table properties (object DataProperties).",
    )
    parser.add_argument(
        "--update",
        "-u",
        dest="data_refresh",
        default=False,
        action="store_true",
        help="Will update previously downloaded data in data folders.",
    )
    parser.add_argument(
        "--force",
        dest="force_download",
        default=False,
        action="store_true",
        help="When downloading data will remove old data files",
    )
    parser.add_argument(
        "--modified",
        "-w",
        metavar='',
        type=str,
        dest="modified_within",
        default="",
        nargs="?",
        help=argparse.SUPPRESS,
    )
    parser.add_argument(
        "--downloadcsv",
        "-c",
        dest="download_csv",
        default=False,
        action="store_true",
        help="Download CSV files.",
    )
    parser.add_argument(
        "--downloadexcel",
        "-x",
        dest="download_excel",
        default=False,
        action="store_true",
        help="Create excel files from downloaded data.",
    )
    parser.add_argument(
        "--downloadjson",
        "-j",
        dest="download_json",
        default=False,
        action="store_true",
        help="Download Json data files.",
    )

    global settings
    global silent
    global critical
    global error
    global warning
    global info
    global verbose
    global allmsg

    options = vars(parser.parse_args())
    settings = get_defaults()
    settings.update(options)
    #  no arguments given, update silently
    if len(sys.argv)==1: 
        settings["data_refresh"] = True
        settings["loglevel"] = "critical"

    #  who we are and what we do
    p(verbose,
        str(sys.argv).replace('[','').replace(']','').replace("'",'')+'\n')

    #  general download settings (default is all)
    if settings["download_data"] and (
            not settings["download_csv"] and
            not settings["download_excel"] and
            not settings["download_json"]):
        settings["download_csv"] = True
        settings["download_excel"] = True
        settings["download_json"] = True
    
    #  must have -d when having c, x or j
    if not settings["download_data"] and \
                not settings["data_refresh"] and (
                settings["download_csv"]
                or settings["download_excel"] 
                or settings["download_json"]):
        download_data = True

    #  only update local master when actually downloading data
    if not settings["download_data"] and not settings["data_refresh"]: 
        settings["no_master"] = True

    #  message
    p(verbose, [['-->' if type(
            p(verbose, i[0] + '\t\t= ' + str(i[1]) if i else '')) == None 
            else 'Not'] for i in settings.items()])

    # get table list
    getTableList()

#  ********
#  start
if __name__ == "__main__":
    #  --------
    #  timing program
    begin = datetime.datetime.now()

    #  --------
    #  control information
    controlInformationTables = {}
    controlInformationTable = {
        "Title":"","Updated":"","ShortTitle":"","Identifier":"",
        "Summary":"","Modified":"","ReasonDelivery":"",
        "Frequency":"","Period":"","RecordCount":"","lastRefreshDate":"",
        "lastRefreshDateJson":"","lastRefreshDateCsv":"",
        "lastRefreshDateExcel":""}
    control_file = 'get_data_control.xlsx'

    #  --------
    #  loglevels CONSTANTS
    loglevels = ["silent","critical","error","warning",
                 "info","verbose","allmsg"]
    silent = 0
    critical = 1
    error = 2
    warning = 3
    info = 4
    verbose = 5
    allmsg = 6

    main()
